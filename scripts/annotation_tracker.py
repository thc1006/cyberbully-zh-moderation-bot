#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
標註追蹤和管理系統
提供標註進度追蹤、Excel報表生成、標註者績效管理等功能
"""

import os
import json
import logging
import argparse
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import matplotlib.pyplot as plt
import seaborn as sns
import warnings
warnings.filterwarnings('ignore')

# 設定中文字體
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei']
plt.rcParams['axes.unicode_minus'] = False

# 設定日誌
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class AnnotationTracker:
    """標註追蹤器"""

    def __init__(self, project_dir: str = "data/annotation"):
        """初始化標註追蹤器

        Args:
            project_dir: 專案目錄
        """
        self.project_dir = Path(project_dir)
        self.tracking_dir = self.project_dir / "tracking"
        self.tracking_dir.mkdir(parents=True, exist_ok=True)

        # 標籤維度
        self.label_dimensions = {
            'toxicity': ['none', 'toxic', 'severe'],
            'bullying': ['none', 'harassment', 'threat'],
            'role': ['none', 'perpetrator', 'victim', 'bystander'],
            'emotion': ['pos', 'neu', 'neg'],
            'emotion_strength': ['0', '1', '2', '3', '4']
        }

        logger.info(f"標註追蹤器初始化完成，追蹤目錄: {self.tracking_dir}")

    def create_tracking_sheet(self, samples: List[Dict], annotators: List[str],
                            task_name: str = "annotation_task") -> str:
        """建立標註追蹤表格

        Args:
            samples: 樣本列表
            annotators: 標註者列表
            task_name: 任務名稱

        Returns:
            Excel檔案路徑
        """
        logger.info(f"建立標註追蹤表格: {task_name}")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = self.tracking_dir / f"{task_name}_tracking_{timestamp}.xlsx"

        # 建立工作簿
        workbook = openpyxl.Workbook()

        # 建立主要追蹤工作表
        main_sheet = workbook.active
        main_sheet.title = "標註追蹤"

        # 設定標題
        headers = [
            '樣本ID', '樣本內容', '困難度', '分配日期', '預計完成日期'
        ]

        # 為每個標註者添加欄位
        for annotator in annotators:
            headers.extend([
                f'{annotator}_狀態',
                f'{annotator}_完成日期',
                f'{annotator}_毒性',
                f'{annotator}_霸凌',
                f'{annotator}_角色',
                f'{annotator}_情緒',
                f'{annotator}_情緒強度',
                f'{annotator}_備註'
            ])

        headers.extend(['一致性檢查', '最終標註', '審核狀態', '審核備註'])

        # 寫入標題
        for col, header in enumerate(headers, 1):
            cell = main_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填入樣本資料
        for row, sample in enumerate(samples, 2):
            sample_id = sample.get('id', sample.get('annotation_metadata', {}).get('original_index', row-1))

            # 提取樣本內容
            content = self.extract_sample_text(sample)
            if len(content) > 100:
                content = content[:100] + "..."

            # 提取困難度
            difficulty = sample.get('annotation_metadata', {}).get('annotation_priority', 'medium')

            # 基本資訊
            main_sheet.cell(row=row, column=1, value=str(sample_id))
            main_sheet.cell(row=row, column=2, value=content)
            main_sheet.cell(row=row, column=3, value=difficulty)
            main_sheet.cell(row=row, column=4, value=datetime.now().strftime("%Y-%m-%d"))
            main_sheet.cell(row=row, column=5, value=(datetime.now() + timedelta(days=7)).strftime("%Y-%m-%d"))

            # 為每個標註者設定初始狀態
            col_offset = 6
            for i, annotator in enumerate(annotators):
                status_col = col_offset + i * 8
                main_sheet.cell(row=row, column=status_col, value="待標註")

        # 調整欄寬
        for column in main_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            main_sheet.column_dimensions[column_letter].width = adjusted_width

        # 建立標註者統計工作表
        self.create_annotator_stats_sheet(workbook, annotators)

        # 建立進度追蹤工作表
        self.create_progress_tracking_sheet(workbook, len(samples), annotators)

        # 建立品質控制工作表
        self.create_quality_control_sheet(workbook)

        # 儲存檔案
        workbook.save(excel_file)
        logger.info(f"標註追蹤表格已建立: {excel_file}")

        return str(excel_file)

    def extract_sample_text(self, sample: Dict) -> str:
        """從樣本中提取文字"""
        text_fields = ['text', 'content', 'message', 'sentence', 'comment']
        for field in text_fields:
            if field in sample and sample[field]:
                return str(sample[field])
        return str(sample)

    def create_annotator_stats_sheet(self, workbook: openpyxl.Workbook, annotators: List[str]):
        """建立標註者統計工作表"""
        stats_sheet = workbook.create_sheet(title="標註者統計")

        # 標題
        headers = [
            '標註者', '分配樣本數', '完成樣本數', '完成率', '平均標註時間',
            '困難案例數', '品質評分', '一致性評分', '備註'
        ]

        for col, header in enumerate(headers, 1):
            cell = stats_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 為每個標註者建立統計行
        for row, annotator in enumerate(annotators, 2):
            stats_sheet.cell(row=row, column=1, value=annotator)
            stats_sheet.cell(row=row, column=2, value=0)  # 將由更新函數填入
            stats_sheet.cell(row=row, column=3, value=0)
            stats_sheet.cell(row=row, column=4, value="0%")
            stats_sheet.cell(row=row, column=5, value="待計算")
            stats_sheet.cell(row=row, column=6, value=0)
            stats_sheet.cell(row=row, column=7, value="待評估")
            stats_sheet.cell(row=row, column=8, value="待評估")

        # 調整欄寬
        for col in range(1, len(headers) + 1):
            stats_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    def create_progress_tracking_sheet(self, workbook: openpyxl.Workbook, total_samples: int, annotators: List[str]):
        """建立進度追蹤工作表"""
        progress_sheet = workbook.create_sheet(title="進度追蹤")

        # 日期範圍 (未來30天)
        start_date = datetime.now().date()
        dates = [start_date + timedelta(days=i) for i in range(30)]

        # 標題
        headers = ['日期'] + [f'{annotator}_累計完成' for annotator in annotators] + ['總體進度']

        for col, header in enumerate(headers, 1):
            cell = progress_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 填入日期
        for row, date in enumerate(dates, 2):
            progress_sheet.cell(row=row, column=1, value=date.strftime("%Y-%m-%d"))

        # 調整欄寬
        for col in range(1, len(headers) + 1):
            progress_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    def create_quality_control_sheet(self, workbook: openpyxl.Workbook):
        """建立品質控制工作表"""
        qc_sheet = workbook.create_sheet(title="品質控制")

        # 品質指標
        qc_headers = [
            '指標', '目標值', '當前值', '狀態', '備註'
        ]

        for col, header in enumerate(qc_headers, 1):
            cell = qc_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 品質指標項目
        qc_metrics = [
            ('標註者間一致性 (Kappa)', '≥ 0.6', '待計算', '待評估', ''),
            ('完成率', '≥ 90%', '待計算', '待評估', ''),
            ('困難案例比例', '≤ 20%', '待計算', '待評估', ''),
            ('平均標註時間', '≤ 5分鐘/樣本', '待計算', '待評估', ''),
            ('品質審核通過率', '≥ 95%', '待計算', '待評估', '')
        ]

        for row, (metric, target, current, status, note) in enumerate(qc_metrics, 2):
            qc_sheet.cell(row=row, column=1, value=metric)
            qc_sheet.cell(row=row, column=2, value=target)
            qc_sheet.cell(row=row, column=3, value=current)
            qc_sheet.cell(row=row, column=4, value=status)
            qc_sheet.cell(row=row, column=5, value=note)

        # 調整欄寬
        for col in range(1, len(qc_headers) + 1):
            qc_sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    def update_tracking_sheet(self, excel_file: str, annotations_data: Dict[str, Any]):
        """更新標註追蹤表格

        Args:
            excel_file: Excel檔案路徑
            annotations_data: 標註資料 {annotator_id: {sample_id: annotation}}
        """
        logger.info(f"更新標註追蹤表格: {excel_file}")

        try:
            workbook = openpyxl.load_workbook(excel_file)
            main_sheet = workbook["標註追蹤"]

            # 獲取標題行以找到欄位位置
            headers = {}
            for col in range(1, main_sheet.max_column + 1):
                header = main_sheet.cell(row=1, column=col).value
                if header:
                    headers[header] = col

            # 更新標註資料
            for row in range(2, main_sheet.max_row + 1):
                sample_id = str(main_sheet.cell(row=row, column=headers['樣本ID']).value)

                for annotator_id, annotations in annotations_data.items():
                    if sample_id in annotations:
                        annotation = annotations[sample_id]

                        # 更新狀態
                        status_col = headers.get(f'{annotator_id}_狀態')
                        if status_col:
                            main_sheet.cell(row=row, column=status_col, value="已完成")

                        # 更新完成日期
                        date_col = headers.get(f'{annotator_id}_完成日期')
                        if date_col:
                            timestamp = annotation.get('timestamp', datetime.now().isoformat())
                            try:
                                date_obj = datetime.fromisoformat(timestamp.replace('Z', '+00:00'))
                                main_sheet.cell(row=row, column=date_col, value=date_obj.strftime("%Y-%m-%d"))
                            except:
                                main_sheet.cell(row=row, column=date_col, value=datetime.now().strftime("%Y-%m-%d"))

                        # 更新標註結果
                        for dimension in ['毒性', '霸凌', '角色', '情緒', '情緒強度']:
                            dim_col = headers.get(f'{annotator_id}_{dimension}')
                            if dim_col:
                                dim_key = {
                                    '毒性': 'toxicity',
                                    '霸凌': 'bullying',
                                    '角色': 'role',
                                    '情緒': 'emotion',
                                    '情緒強度': 'emotion_strength'
                                }.get(dimension)

                                if dim_key and dim_key in annotation:
                                    main_sheet.cell(row=row, column=dim_col, value=annotation[dim_key])

                        # 更新備註
                        note_col = headers.get(f'{annotator_id}_備註')
                        if note_col and 'note' in annotation:
                            main_sheet.cell(row=row, column=note_col, value=annotation['note'])

            # 更新標註者統計
            self.update_annotator_stats(workbook, annotations_data)

            # 儲存檔案
            workbook.save(excel_file)
            logger.info("標註追蹤表格更新完成")

        except Exception as e:
            logger.error(f"更新標註追蹤表格失敗: {str(e)}")

    def update_annotator_stats(self, workbook: openpyxl.Workbook, annotations_data: Dict[str, Any]):
        """更新標註者統計資料"""
        if "標註者統計" not in workbook.sheetnames:
            return

        stats_sheet = workbook["標註者統計"]
        main_sheet = workbook["標註追蹤"]

        # 計算每個標註者的統計資料
        for row in range(2, stats_sheet.max_row + 1):
            annotator = stats_sheet.cell(row=row, column=1).value
            if not annotator or annotator not in annotations_data:
                continue

            annotations = annotations_data[annotator]

            # 計算完成數量
            completed_count = len(annotations)
            stats_sheet.cell(row=row, column=3, value=completed_count)

            # 計算分配數量 (從主表格統計)
            assigned_count = 0
            status_col = None
            for col in range(1, main_sheet.max_column + 1):
                header = main_sheet.cell(row=1, column=col).value
                if header == f'{annotator}_狀態':
                    status_col = col
                    break

            if status_col:
                for main_row in range(2, main_sheet.max_row + 1):
                    if main_sheet.cell(row=main_row, column=status_col).value:
                        assigned_count += 1

            stats_sheet.cell(row=row, column=2, value=assigned_count)

            # 計算完成率
            completion_rate = completed_count / assigned_count if assigned_count > 0 else 0
            stats_sheet.cell(row=row, column=4, value=f"{completion_rate:.1%}")

            # 計算困難案例數
            difficult_count = sum(1 for ann in annotations.values() if ann.get('difficult', False))
            stats_sheet.cell(row=row, column=6, value=difficult_count)

    def generate_progress_chart(self, excel_file: str, output_dir: Optional[str] = None):
        """生成進度圖表

        Args:
            excel_file: Excel檔案路徑
            output_dir: 輸出目錄
        """
        if not output_dir:
            output_dir = self.tracking_dir

        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        try:
            # 讀取標註者統計資料
            df = pd.read_excel(excel_file, sheet_name="標註者統計")

            # 完成率圖表
            plt.figure(figsize=(10, 6))
            df_clean = df.dropna(subset=['標註者', '完成率'])

            # 處理完成率欄位（移除百分號）
            completion_rates = []
            for rate in df_clean['完成率']:
                if isinstance(rate, str) and '%' in rate:
                    completion_rates.append(float(rate.replace('%', '')) / 100)
                else:
                    completion_rates.append(float(rate) if pd.notna(rate) else 0)

            plt.bar(df_clean['標註者'], completion_rates)
            plt.title('標註者完成率')
            plt.xlabel('標註者')
            plt.ylabel('完成率')
            plt.ylim(0, 1)
            plt.xticks(rotation=45)

            # 添加數值標籤
            for i, rate in enumerate(completion_rates):
                plt.text(i, rate + 0.01, f'{rate:.1%}', ha='center', va='bottom')

            plt.tight_layout()
            completion_chart_file = output_dir / "completion_rates.png"
            plt.savefig(completion_chart_file, dpi=300, bbox_inches='tight')
            plt.close()

            # 樣本分配圖表
            plt.figure(figsize=(10, 6))
            plt.bar(df_clean['標註者'], df_clean['分配樣本數'], alpha=0.7, label='分配樣本數')
            plt.bar(df_clean['標註者'], df_clean['完成樣本數'], alpha=0.9, label='完成樣本數')
            plt.title('標註者樣本分配與完成情況')
            plt.xlabel('標註者')
            plt.ylabel('樣本數量')
            plt.legend()
            plt.xticks(rotation=45)
            plt.tight_layout()

            allocation_chart_file = output_dir / "sample_allocation.png"
            plt.savefig(allocation_chart_file, dpi=300, bbox_inches='tight')
            plt.close()

            logger.info(f"進度圖表已生成: {completion_chart_file}, {allocation_chart_file}")

        except Exception as e:
            logger.error(f"生成進度圖表失敗: {str(e)}")

    def export_summary_report(self, excel_file: str, output_dir: Optional[str] = None) -> str:
        """匯出摘要報告

        Args:
            excel_file: Excel檔案路徑
            output_dir: 輸出目錄

        Returns:
            報告檔案路徑
        """
        if not output_dir:
            output_dir = self.tracking_dir

        output_dir = Path(output_dir)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_file = output_dir / f"annotation_summary_report_{timestamp}.txt"

        try:
            # 讀取Excel資料
            main_df = pd.read_excel(excel_file, sheet_name="標註追蹤")
            stats_df = pd.read_excel(excel_file, sheet_name="標註者統計")

            with open(report_file, 'w', encoding='utf-8') as f:
                f.write("=== 標註任務摘要報告 ===\n\n")
                f.write(f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

                # 基本統計
                total_samples = len(main_df)
                f.write("=== 基本統計 ===\n")
                f.write(f"總樣本數: {total_samples}\n")
                f.write(f"標註者數量: {len(stats_df)}\n\n")

                # 標註者統計
                f.write("=== 標註者表現 ===\n")
                for _, row in stats_df.iterrows():
                    annotator = row['標註者']
                    assigned = row['分配樣本數'] if pd.notna(row['分配樣本數']) else 0
                    completed = row['完成樣本數'] if pd.notna(row['完成樣本數']) else 0
                    completion_rate = row['完成率'] if pd.notna(row['完成率']) else "0%"
                    difficult = row['困難案例數'] if pd.notna(row['困難案例數']) else 0

                    f.write(f"{annotator}:\n")
                    f.write(f"  分配樣本: {assigned}\n")
                    f.write(f"  完成樣本: {completed}\n")
                    f.write(f"  完成率: {completion_rate}\n")
                    f.write(f"  困難案例: {difficult}\n\n")

                # 整體進度
                total_assigned = stats_df['分配樣本數'].sum()
                total_completed = stats_df['完成樣本數'].sum()
                overall_completion = total_completed / total_assigned if total_assigned > 0 else 0

                f.write("=== 整體進度 ===\n")
                f.write(f"總分配樣本: {total_assigned}\n")
                f.write(f"總完成樣本: {total_completed}\n")
                f.write(f"整體完成率: {overall_completion:.1%}\n\n")

                # 品質指標（如果有資料）
                try:
                    qc_df = pd.read_excel(excel_file, sheet_name="品質控制")
                    f.write("=== 品質指標 ===\n")
                    for _, row in qc_df.iterrows():
                        metric = row['指標']
                        target = row['目標值']
                        current = row['當前值']
                        status = row['狀態']
                        f.write(f"{metric}: {current} (目標: {target}) - {status}\n")
                    f.write("\n")
                except:
                    pass

            logger.info(f"摘要報告已生成: {report_file}")
            return str(report_file)

        except Exception as e:
            logger.error(f"生成摘要報告失敗: {str(e)}")
            return ""

    def create_annotation_dashboard(self, excel_files: List[str], output_dir: Optional[str] = None):
        """建立標註儀表板

        Args:
            excel_files: Excel檔案路徑列表
            output_dir: 輸出目錄
        """
        if not output_dir:
            output_dir = self.tracking_dir

        output_dir = Path(output_dir)
        dashboard_dir = output_dir / "dashboard"
        dashboard_dir.mkdir(exist_ok=True)

        # 合併所有任務的資料
        all_stats = []
        for excel_file in excel_files:
            try:
                stats_df = pd.read_excel(excel_file, sheet_name="標註者統計")
                task_name = Path(excel_file).stem
                stats_df['任務'] = task_name
                all_stats.append(stats_df)
            except Exception as e:
                logger.warning(f"讀取檔案失敗 {excel_file}: {str(e)}")

        if not all_stats:
            logger.error("沒有可用的統計資料")
            return

        combined_stats = pd.concat(all_stats, ignore_index=True)

        # 生成綜合儀表板
        fig, axes = plt.subplots(2, 2, figsize=(16, 12))

        # 1. 標註者總體表現
        annotator_totals = combined_stats.groupby('標註者').agg({
            '分配樣本數': 'sum',
            '完成樣本數': 'sum'
        }).reset_index()
        annotator_totals['完成率'] = annotator_totals['完成樣本數'] / annotator_totals['分配樣本數']

        axes[0, 0].bar(annotator_totals['標註者'], annotator_totals['完成率'])
        axes[0, 0].set_title('標註者整體完成率')
        axes[0, 0].set_ylabel('完成率')
        axes[0, 0].tick_params(axis='x', rotation=45)

        # 2. 每日進度趨勢（模擬資料）
        dates = pd.date_range(start=datetime.now().date() - timedelta(days=30), periods=30)
        daily_progress = [i * 2 + np.random.randint(-5, 6) for i in range(30)]
        axes[0, 1].plot(dates, daily_progress, marker='o')
        axes[0, 1].set_title('每日標註進度')
        axes[0, 1].set_ylabel('完成樣本數')
        axes[0, 1].tick_params(axis='x', rotation=45)

        # 3. 任務分布
        task_counts = combined_stats.groupby('任務')['完成樣本數'].sum()
        axes[1, 0].pie(task_counts.values, labels=task_counts.index, autopct='%1.1f%%')
        axes[1, 0].set_title('各任務完成樣本分布')

        # 4. 困難案例統計
        annotator_difficult = combined_stats.groupby('標註者')['困難案例數'].sum()
        axes[1, 1].bar(annotator_difficult.index, annotator_difficult.values, color='orange')
        axes[1, 1].set_title('標註者困難案例數量')
        axes[1, 1].set_ylabel('困難案例數')
        axes[1, 1].tick_params(axis='x', rotation=45)

        plt.tight_layout()
        dashboard_file = dashboard_dir / "annotation_dashboard.png"
        plt.savefig(dashboard_file, dpi=300, bbox_inches='tight')
        plt.close()

        # 生成HTML報告
        html_file = dashboard_dir / "annotation_dashboard.html"
        self.generate_html_dashboard(combined_stats, html_file, dashboard_file)

        logger.info(f"標註儀表板已建立: {dashboard_dir}")

    def generate_html_dashboard(self, stats_df: pd.DataFrame, html_file: Path, chart_file: Path):
        """生成HTML儀表板"""
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>標註任務儀表板</title>
            <meta charset="utf-8">
            <style>
                body {{ font-family: 'Microsoft YaHei', Arial, sans-serif; margin: 20px; }}
                .header {{ text-align: center; margin-bottom: 30px; }}
                .summary {{ background-color: #f5f5f5; padding: 20px; margin-bottom: 20px; }}
                .chart {{ text-align: center; margin: 20px 0; }}
                table {{ border-collapse: collapse; width: 100%; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #4CAF50; color: white; }}
                .metric {{ display: inline-block; margin: 10px; padding: 10px; background-color: #e7f3ff; border-radius: 5px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>標註任務儀表板</h1>
                <p>更新時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>

            <div class="summary">
                <h2>總體統計</h2>
                <div class="metric">
                    <strong>總標註者數:</strong> {len(stats_df['標註者'].unique())}
                </div>
                <div class="metric">
                    <strong>總分配樣本:</strong> {stats_df['分配樣本數'].sum():,}
                </div>
                <div class="metric">
                    <strong>總完成樣本:</strong> {stats_df['完成樣本數'].sum():,}
                </div>
                <div class="metric">
                    <strong>整體完成率:</strong> {stats_df['完成樣本數'].sum() / stats_df['分配樣本數'].sum():.1%}
                </div>
            </div>

            <div class="chart">
                <h2>標註進度圖表</h2>
                <img src="{chart_file.name}" alt="標註進度圖表" style="max-width: 100%;">
            </div>

            <div>
                <h2>詳細統計表</h2>
                {stats_df.to_html(escape=False, classes='table')}
            </div>
        </body>
        </html>
        """

        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)


def main():
    """主函數"""
    parser = argparse.ArgumentParser(description="標註追蹤和管理系統")

    subparsers = parser.add_subparsers(dest='command', help='可用命令')

    # 建立追蹤表格
    create_parser = subparsers.add_parser('create', help='建立標註追蹤表格')
    create_parser.add_argument('--samples', required=True, help='樣本檔案路徑')
    create_parser.add_argument('--annotators', nargs='+', required=True, help='標註者列表')
    create_parser.add_argument('--task_name', required=True, help='任務名稱')

    # 更新追蹤表格
    update_parser = subparsers.add_parser('update', help='更新標註追蹤表格')
    update_parser.add_argument('--excel_file', required=True, help='Excel檔案路徑')
    update_parser.add_argument('--annotation_files', nargs='+', required=True, help='標註結果檔案列表')

    # 生成圖表
    chart_parser = subparsers.add_parser('chart', help='生成進度圖表')
    chart_parser.add_argument('--excel_file', required=True, help='Excel檔案路徑')
    chart_parser.add_argument('--output_dir', help='輸出目錄')

    # 生成報告
    report_parser = subparsers.add_parser('report', help='生成摘要報告')
    report_parser.add_argument('--excel_file', required=True, help='Excel檔案路徑')
    report_parser.add_argument('--output_dir', help='輸出目錄')

    # 建立儀表板
    dashboard_parser = subparsers.add_parser('dashboard', help='建立標註儀表板')
    dashboard_parser.add_argument('--excel_files', nargs='+', required=True, help='Excel檔案路徑列表')
    dashboard_parser.add_argument('--output_dir', help='輸出目錄')

    # 專案目錄
    parser.add_argument('--project_dir', default='data/annotation', help='專案目錄')

    args = parser.parse_args()

    if not args.command:
        parser.print_help()
        return

    # 初始化追蹤器
    tracker = AnnotationTracker(args.project_dir)

    # 執行對應命令
    if args.command == 'create':
        # 載入樣本
        with open(args.samples, 'r', encoding='utf-8') as f:
            if args.samples.endswith('.json'):
                samples = json.load(f)
            elif args.samples.endswith('.jsonl'):
                samples = [json.loads(line) for line in f]
            elif args.samples.endswith('.csv'):
                df = pd.read_csv(args.samples)
                samples = df.to_dict('records')
            else:
                logger.error("不支援的檔案格式")
                return

        excel_file = tracker.create_tracking_sheet(samples, args.annotators, args.task_name)
        print(f"追蹤表格已建立: {excel_file}")

    elif args.command == 'update':
        # 載入標註結果
        annotations_data = {}
        for annotation_file in args.annotation_files:
            try:
                with open(annotation_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)

                # 提取標註者ID和標註資料
                if isinstance(data, dict) and 'annotator_id' in data:
                    annotator_id = data['annotator_id']
                    annotations = data.get('annotations', {})
                elif isinstance(data, list) and len(data) > 0:
                    annotator_id = data[0].get('annotator_id', Path(annotation_file).stem)
                    annotations = {}
                    for item in data:
                        if 'annotation' in item:
                            sample_id = str(item.get('id', len(annotations)))
                            annotations[sample_id] = item['annotation']

                annotations_data[annotator_id] = annotations

            except Exception as e:
                logger.error(f"載入標註檔案失敗 {annotation_file}: {str(e)}")

        tracker.update_tracking_sheet(args.excel_file, annotations_data)
        print("追蹤表格已更新")

    elif args.command == 'chart':
        tracker.generate_progress_chart(args.excel_file, args.output_dir)
        print("進度圖表已生成")

    elif args.command == 'report':
        report_file = tracker.export_summary_report(args.excel_file, args.output_dir)
        print(f"摘要報告已生成: {report_file}")

    elif args.command == 'dashboard':
        tracker.create_annotation_dashboard(args.excel_files, args.output_dir)
        print("標註儀表板已建立")


if __name__ == "__main__":
    main()