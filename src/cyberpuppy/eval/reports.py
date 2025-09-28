"""
報告生成模組
提供多格式的評估報告生成功能
"""

import json
import logging
import os
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional

import numpy as np
import pandas as pd

# PDF 生成
try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, letter
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate,
                                    Spacer, Table, TableStyle)

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab 未安裝，PDF 功能將不可用")

# Excel 生成
try:
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl 未安裝，Excel 功能將不可用")

logger = logging.getLogger(__name__)


@dataclass
class ReportConfig:
    """報告配置"""

    title: str = "CyberPuppy 霸凌偵測系統評估報告"
    subtitle: str = "模型效能與分析報告"
    author: str = "CyberPuppy Team"
    company: str = "網路霸凌防治系統"
    logo_path: Optional[str] = None
    include_charts: bool = True
    include_detailed_analysis: bool = True
    language: str = "zh-TW"


class ReportGenerator:
    """報告生成器主類"""

    def __init__(self, output_dir: str = "reports"):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

        # 設定報告樣式
        self.colors = {
            "primary": "#2E86AB",
            "secondary": "#A23B72",
            "success": "#28A745",
            "warning": "#FFC107",
            "danger": "#DC3545",
            "info": "#17A2B8",
        }

    def generate_comprehensive_report(
        self,
        evaluation_results: Dict[str, Any],
        error_analysis: Dict[str, Any] = None,
        robustness_results: Dict[str, Any] = None,
        explanations: List[Dict[str, Any]] = None,
        config: ReportConfig = None,
        formats: List[str] = None,
    ) -> Dict[str, str]:
        """
        生成綜合評估報告（多格式）

        Args:
            evaluation_results: 評估結果
            error_analysis: 錯誤分析結果
            robustness_results: 穩健性測試結果
            explanations: 可解釋性分析結果
            config: 報告配置
            formats: 輸出格式列表

        Returns:
            生成的報告文件路徑字典
        """

        if formats is None:
            formats = ["html", "pdf", "json", "excel"]
        if config is None:
            config = ReportConfig()

        logger.info(f"開始生成綜合報告，格式: {formats}")

        # 整合所有數據
        report_data = {
            "metadata": {
                "title": config.title,
                "subtitle": config.subtitle,
                "author": config.author,
                "company": config.company,
                "generation_time": datetime.now().isoformat(),
                "version": "1.0.0",
            },
            "evaluation_results": evaluation_results,
            "error_analysis": error_analysis,
            "robustness_results": robustness_results,
            "explanations": explanations,
        }

        generated_files = {}

        # 生成各種格式的報告
        for format_type in formats:
            try:
                if format_type.lower() == "html":
                    filepath = self._generate_html_report(report_data, config)
                    generated_files["html"] = filepath
                elif format_type.lower() == "pdf":
                    filepath = self._generate_pdf_report(report_data, config)
                    generated_files["pdf"] = filepath
                elif format_type.lower() == "json":
                    filepath = self._generate_json_report(report_data, config)
                    generated_files["json"] = filepath
                elif format_type.lower() == "excel":
                    filepath = self._generate_excel_report(report_data, config)
                    generated_files["excel"] = filepath
                else:
                    logger.warning(f"不支援的報告格式: {format_type}")

            except Exception as e:
                logger.error(f"生成 {format_type} 報告時發生錯誤: {str(e)}")

        logger.info(f"報告生成完成，共生成 {len(generated_files)} 個文件")

        return generated_files

    def _generate_html_report(self, report_data: Dict[str, Any], config: ReportConfig) -> str:
        """生成 HTML 報告"""

        logger.info("生成 HTML 報告...")

        # HTML 模板
        html_template = """
        <!DOCTYPE html>
        <html lang="{language}">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>{title}</title>
            <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
            <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
            <style>
                {css_styles}
            </style>
        </head>
        <body>
            <div class="container">
                {content}
            </div>
            <script>
                {javascript}
            </script>
        </body>
        </html>
        """

        # CSS 樣式
        css_styles = """
        body {
            font-family: 'Arial', 'Microsoft JhengHei', sans-serif;
            margin: 0;
            padding: 0;
            background-color: #f5f5f5;
            color: #333;
        }
        .container {
            max-width: 1200px;
            margin: 0 auto;
            background-color: white;
            min-height: 100vh;
        }
        .header {
            background: linear-gradient(135deg, #2E86AB, #A23B72);
            color: white;
            padding: 40px 30px;
            text-align: center;
        }
        .header h1 {
            margin: 0;
            font-size: 2.5em;
            font-weight: bold;
        }
        .header h2 {
            margin: 10px 0 0 0;
            font-size: 1.5em;
            font-weight: normal;
            opacity: 0.9;
        }
        .meta-info {
            background-color: #f8f9fa;
            padding: 20px 30px;
            border-bottom: 1px solid #ddd;
        }
        .meta-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
        }
        .meta-item {
            display: flex;
            justify-content: space-between;
        }
        .meta-label {
            font-weight: bold;
            color: #666;
        }
        .section {
            padding: 30px;
            border-bottom: 1px solid #eee;
        }
        .section h2 {
            color: #2E86AB;
            border-bottom: 2px solid #2E86AB;
            padding-bottom: 10px;
            margin-bottom: 25px;
        }
        .metrics-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            margin: 25px 0;
        }
        .metric-card {
            background: linear-gradient(135deg, #ffffff, #f8f9fa);
            padding: 25px;
            border-radius: 10px;
            text-align: center;
            box-shadow: 0 4px 6px rgba(0,0,0,0.1);
            border: 1px solid #e3e6ea;
        }
        .metric-value {
            font-size: 2.5em;
            font-weight: bold;
            color: #2E86AB;
            margin-bottom: 5px;
        }
        .metric-label {
            color: #666;
            font-weight: 500;
            text-transform: uppercase;
            font-size: 0.9em;
        }
        .chart-container {
            width: 100%;
            height: 400px;
            margin: 25px 0;
            background: white;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            padding: 20px;
        }
        .table-container {
            overflow-x: auto;
            margin: 25px 0;
        }
        .data-table {
            width: 100%;
            border-collapse: collapse;
            background: white;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            border-radius: 8px;
            overflow: hidden;
        }
        .data-table th {
            background-color: #2E86AB;
            color: white;
            padding: 15px;
            text-align: left;
            font-weight: bold;
        }
        .data-table td {
            padding: 12px 15px;
            border-bottom: 1px solid #eee;
        }
        .data-table tr:hover {
            background-color: #f8f9fa;
        }
        .status-badge {
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 0.85em;
            font-weight: bold;
            text-transform: uppercase;
        }
        .status-excellent {
            background-color: #d4edda;
            color: #155724;
        }
        .status-good {
            background-color: #d1ecf1;
            color: #0c5460;
        }
        .status-warning {
            background-color: #fff3cd;
            color: #856404;
        }
        .status-poor {
            background-color: #f8d7da;
            color: #721c24;
        }
        .tabs {
            display: flex;
            background-color: #f1f1f1;
            border-radius: 8px 8px 0 0;
            overflow: hidden;
        }
        .tab {
            padding: 15px 25px;
            cursor: pointer;
            border: none;
            background-color: transparent;
            flex: 1;
            text-align: center;
            font-weight: bold;
            transition: all 0.3s ease;
        }
        .tab:hover {
            background-color: #e9ecef;
        }
        .tab.active {
            background-color: #2E86AB;
            color: white;
        }
        .tab-content {
            display: none;
            padding: 30px;
            border: 1px solid #ddd;
            border-top: none;
            border-radius: 0 0 8px 8px;
            background: white;
        }
        .tab-content.active {
            display: block;
        }
        .recommendation-list {
            list-style: none;
            padding: 0;
        }
        .recommendation-item {
            background: #f8f9fa;
            margin: 10px 0;
            padding: 15px;
            border-left: 4px solid #2E86AB;
            border-radius: 0 8px 8px 0;
        }
        .footer {
            background-color: #2c3e50;
            color: white;
            text-align: center;
            padding: 30px;
        }
        """

        # 生成內容
        content_sections = []

        # 頁首
        header_content = f"""
        <div class="header">
            <h1>{config.title}</h1>
            <h2>{config.subtitle}</h2>
        </div>
        """
        content_sections.append(header_content)

        # 元數據信息
        metadata = report_data["metadata"]
        meta_content = f"""
        <div class="meta-info">
            <div class="meta-grid">
                <div class="meta-item">
                    <span class="meta-label">作者:</span>
                    <span>{metadata['author']}</span>
                </div>
                <div class="meta-item">
                    <span class="meta-label">公司:</span>
                    <span>{metadata['company']}</span>
                </div>
                <div class="meta-item">
                    <span class="meta-label">生成時間:</span>
                    <span>{datetime.fromisoformat(metadata['generation_time']).strftime('%Y-%m-%d %H:%M:%S')}</span>
                </div>
                <div class="meta-item">
                    <span class="meta-label">版本:</span>
                    <span>{metadata['version']}</span>
                </div>
            </div>
        </div>
        """
        content_sections.append(meta_content)

        # 總體指標
        if report_data["evaluation_results"] and "metrics" in report_data["evaluation_results"]:
            metrics = report_data["evaluation_results"]["metrics"]
            metrics_content = f"""
            <div class="section">
                <h2>總體效能指標</h2>
                <div class="metrics-grid">
                    <div class="metric-card">
                        <div class="metric-value">{metrics.get('precision', 0):.3f}</div>
                        <div class="metric-label">Precision</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{metrics.get('recall', 0):.3f}</div>
                        <div class="metric-label">Recall</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{metrics.get('f1', 0):.3f}</div>
                        <div class="metric-label">F1-Score</div>
                    </div>
                    <div class="metric-card">
                        <div class="metric-value">{metrics.get('accuracy', 0):.3f}</div>
                        <div class="metric-label">Accuracy</div>
                    </div>
                </div>
            </div>
            """
            content_sections.append(metrics_content)

        # 詳細分析標籤頁
        tabs_content = """
        <div class="section">
            <div class="tabs">
                <button class="tab active" onclick="showTab('performance', this)">效能分析</button>
                <button class="tab" onclick="showTab('errors', this)">錯誤分析</button>
                <button class="tab" onclick="showTab('robustness', this)">穩健性測試</button>
                <button class="tab" onclick="showTab('recommendations', this)">改進建議</button>
            </div>

            <div id="performance" class="tab-content active">
                <h3>效能分析詳情</h3>
                {performance_content}
            </div>

            <div id="errors" class="tab-content">
                <h3>錯誤分析詳情</h3>
                {error_content}
            </div>

            <div id="robustness" class="tab-content">
                <h3>穩健性測試詳情</h3>
                {robustness_content}
            </div>

            <div id="recommendations" class="tab-content">
                <h3>改進建議</h3>
                {recommendations_content}
            </div>
        </div>
        """

        # 填充各標籤頁內容
        performance_content = self._generate_performance_content(report_data["evaluation_results"])
        error_content = self._generate_error_content(report_data["error_analysis"])
        robustness_content = self._generate_robustness_content(report_data["robustness_results"])
        recommendations_content = self._generate_recommendations_content(report_data)

        tabs_content = tabs_content.format(
            performance_content=performance_content,
            error_content=error_content,
            robustness_content=robustness_content,
            recommendations_content=recommendations_content,
        )

        content_sections.append(tabs_content)

        # 頁尾
        footer_content = f"""
        <div class="footer">
            <p>© 2024 {config.company}. 保留所有權利。</p>
            <p>本報告由 CyberPuppy 系統自動生成</p>
        </div>
        """
        content_sections.append(footer_content)

        # JavaScript 功能
        javascript = """
        function showTab(tabName, element) {
            // 隱藏所有標籤內容
            var contents = document.getElementsByClassName('tab-content');
            for (var i = 0; i < contents.length; i++) {
                contents[i].classList.remove('active');
            }

            // 移除所有標籤的 active 類
            var tabs = document.getElementsByClassName('tab');
            for (var i = 0; i < tabs.length; i++) {
                tabs[i].classList.remove('active');
            }

            // 顯示選中的標籤內容
            document.getElementById(tabName).classList.add('active');
            element.classList.add('active');
        }
        """

        # 組合完整內容
        full_content = "\n".join(content_sections)

        # 填入模板
        html_content = html_template.format(
            language=config.language,
            title=config.title,
            css_styles=css_styles,
            content=full_content,
            javascript=javascript,
        )

        # 保存文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"evaluation_report_{timestamp}.html"
        filepath = os.path.join(self.output_dir, filename)

        with open(filepath, "w", encoding="utf-8") as f:
            f.write(html_content)

        logger.info(f"HTML 報告已保存至: {filepath}")
        return filepath

    def _generate_performance_content(self, evaluation_results: Dict[str, Any]) -> str:
        """生成效能分析內容"""

        if not evaluation_results:
            return "<p>無效能分析數據</p>"

        content = []

        # 類別詳細指標
        if "per_class_metrics" in evaluation_results:
            per_class = evaluation_results["per_class_metrics"]

            table_html = """
            <div class="table-container">
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>類別</th>
                            <th>Precision</th>
                            <th>Recall</th>
                            <th>F1-Score</th>
                            <th>Support</th>
                            <th>狀態</th>
                        </tr>
                    </thead>
                    <tbody>
            """

            for class_name, metrics in per_class.items():
                f1_score = metrics.get("f1-score", 0)

                # 確定狀態
                if f1_score >= 0.9:
                    status = '<span class="status-badge status-excellent">優秀</span>'
                elif f1_score >= 0.8:
                    status = '<span class="status-badge status-good">良好</span>'
                elif f1_score >= 0.7:
                    status = '<span class="status-badge status-warning">一般</span>'
                else:
                    status = '<span class="status-badge status-poor">需改進</span>'

                table_html += f"""
                <tr>
                    <td><strong>{class_name}</strong></td>
                    <td>{metrics.get('precision', 0):.3f}</td>
                    <td>{metrics.get('recall', 0):.3f}</td>
                    <td>{metrics.get('f1-score', 0):.3f}</td>
                    <td>{int(metrics.get('support', 0))}</td>
                    <td>{status}</td>
                </tr>
                """

            table_html += """
                    </tbody>
                </table>
            </div>
            """

            content.append(table_html)

        # 混淆矩陣
        if "confusion_matrix" in evaluation_results:
            content.append(
                '<div class="chart-container"><h4>混淆矩陣</h4><p>混淆矩陣圖表將在這裡顯示</p></div>'
            )

        return "\n".join(content)

    def _generate_error_content(self, error_analysis: Dict[str, Any]) -> str:
        """生成錯誤分析內容"""

        if not error_analysis:
            return "<p>無錯誤分析數據</p>"

        content = []

        # 錯誤統計
        if "statistics" in error_analysis:
            stats = error_analysis["statistics"]

            stats_html = f"""
            <div class="metrics-grid">
                <div class="metric-card">
                    <div class="metric-value">{stats.get('total_errors', 0)}</div>
                    <div class="metric-label">總錯誤數</div>
                </div>
            """

            if "error_types" in stats:
                error_types = stats["error_types"]
                for error_type, count in error_types.items():
                    stats_html += f"""
                    <div class="metric-card">
                        <div class="metric-value">{count}</div>
                        <div class="metric-label">{error_type}</div>
                    </div>
                    """

            stats_html += "</div>"
            content.append(stats_html)

        # 改進建議
        if "improvement_suggestions" in error_analysis:
            suggestions = error_analysis["improvement_suggestions"]
            if suggestions:
                suggestions_html = "<h4>改進建議</h4><ul class='recommendation-list'>"
                for suggestion in suggestions[:5]:  # 顯示前5個建議
                    suggestions_html += f"<li class='recommendation-item'>{suggestion}</li>"
                suggestions_html += "</ul>"
                content.append(suggestions_html)

        return "\n".join(content)

    def _generate_robustness_content(self, robustness_results: Dict[str, Any]) -> str:
        """生成穩健性測試內容"""

        if not robustness_results:
            return "<p>無穩健性測試數據</p>"

        content = []

        # 總體穩健性分數
        if "overall_statistics" in robustness_results:
            overall = robustness_results["overall_statistics"]
            score = overall.get("overall_robustness_score", 0)
            level = overall.get("robustness_level", "unknown")

            score_html = f"""
            <div class="metrics-grid">
                <div class="metric-card">
                    <div class="metric-value">{score:.3f}</div>
                    <div class="metric-label">穩健性分數</div>
                </div>
                <div class="metric-card">
                    <div class="metric-value">{level.upper()}</div>
                    <div class="metric-label">穩健性等級</div>
                </div>
            </div>
            """
            content.append(score_html)

        # 攻擊測試結果
        if "summary_statistics" in robustness_results:
            summary = robustness_results["summary_statistics"]

            table_html = """
            <div class="table-container">
                <h4>攻擊測試結果</h4>
                <table class="data-table">
                    <thead>
                        <tr>
                            <th>攻擊類型</th>
                            <th>成功率</th>
                            <th>平均信心下降</th>
                            <th>測試數量</th>
                        </tr>
                    </thead>
                    <tbody>
            """

            for attack_type, stats in summary.items():
                success_rate = stats.get("attack_success_rate", 0)
                conf_drop = stats.get("average_confidence_drop", 0)
                total_tests = stats.get("total_tests", 0)

                table_html += f"""
                <tr>
                    <td>{attack_type.replace('_', ' ').title()}</td>
                    <td>{success_rate:.3f}</td>
                    <td>{conf_drop:.3f}</td>
                    <td>{total_tests}</td>
                </tr>
                """

            table_html += """
                    </tbody>
                </table>
            </div>
            """

            content.append(table_html)

        return "\n".join(content)

    def _generate_recommendations_content(self, report_data: Dict[str, Any]) -> str:
        """生成改進建議內容"""

        recommendations = []

        # 從錯誤分析中獲取建議
        if (
            report_data["error_analysis"]
            and "improvement_suggestions" in report_data["error_analysis"]
        ):
            recommendations.extend(report_data["error_analysis"]["improvement_suggestions"])

        # 從穩健性測試中獲取建議
        if (
            report_data["robustness_results"]
            and "recommendations" in report_data["robustness_results"]
        ):
            recommendations.extend(report_data["robustness_results"]["recommendations"])

        if not recommendations:
            return "<p>暫無改進建議</p>"

        content = "<ul class='recommendation-list'>"
        for i, recommendation in enumerate(recommendations[:10], 1):  # 最多顯示10個建議
            content += (
                f"<li class='recommendation-item'><strong>{i}.</strong> {recommendation}</li>"
            )
        content += "</ul>"

        return content

    def _generate_pdf_report(self, report_data: Dict[str, Any], config: ReportConfig) -> str:
        """生成 PDF 報告"""

        if not REPORTLAB_AVAILABLE:
            logger.error("ReportLab 未安裝，無法生成 PDF 報告")
            return ""

        logger.info("生成 PDF 報告...")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"evaluation_report_{timestamp}.pdf"
        filepath = os.path.join(self.output_dir, filename)

        # 創建 PDF 文檔
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []

        # 獲取樣式
        styles = getSampleStyleSheet()

        # 自定義樣式
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor("#2E86AB"),
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            spaceAfter=20,
            textColor=colors.HexColor("#2E86AB"),
        )

        # 標題
        story.append(Paragraph(config.title, title_style))
        story.append(Paragraph(config.subtitle, styles["Normal"]))
        story.append(Spacer(1, 20))

        # 元數據
        metadata = report_data["metadata"]
        meta_data = [
            ["作者", metadata["author"]],
            ["公司", metadata["company"]],
            [
                "生成時間",
                datetime.fromisoformat(metadata["generation_time"]).strftime("%Y-%m-%d %H:%M:%S"),
            ],
            ["版本", metadata["version"]],
        ]

        meta_table = Table(meta_data, colWidths=[2 * inch, 4 * inch])
        meta_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (0, -1), colors.grey),
                    ("TEXTCOLOR", (0, 0), (0, -1), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                    ("BACKGROUND", (1, 0), (1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(meta_table)
        story.append(Spacer(1, 30))

        # 總體指標
        if report_data["evaluation_results"] and "metrics" in report_data["evaluation_results"]:
            story.append(Paragraph("總體效能指標", heading_style))

            metrics = report_data["evaluation_results"]["metrics"]
            metrics_data = [
                ["指標", "數值"],
                ["Precision", f"{metrics.get('precision', 0):.3f}"],
                ["Recall", f"{metrics.get('recall', 0):.3f}"],
                ["F1-Score", f"{metrics.get('f1', 0):.3f}"],
                ["Accuracy", f"{metrics.get('accuracy', 0):.3f}"],
            ]

            metrics_table = Table(metrics_data, colWidths=[3 * inch, 3 * inch])
            metrics_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )

            story.append(metrics_table)
            story.append(Spacer(1, 20))

        # 類別詳細指標
        if (
            report_data["evaluation_results"]
            and "per_class_metrics" in report_data["evaluation_results"]
        ):

            story.append(Paragraph("類別詳細指標", heading_style))

            per_class = report_data["evaluation_results"]["per_class_metrics"]
            class_data = [["類別", "Precision", "Recall", "F1-Score", "Support"]]

            for class_name, metrics in per_class.items():
                class_data.append(
                    [
                        class_name,
                        f"{metrics.get('precision', 0):.3f}",
                        f"{metrics.get('recall', 0):.3f}",
                        f"{metrics.get('f1-score', 0):.3f}",
                        str(int(metrics.get("support", 0))),
                    ]
                )

            class_table = Table(
                class_data, colWidths=[1.2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch, 1.2 * inch]
            )
            class_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E86AB")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 0), (-1, -1), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )

            story.append(class_table)

        # 建立 PDF
        doc.build(story)

        logger.info(f"PDF 報告已保存至: {filepath}")
        return filepath

    def _generate_json_report(self, report_data: Dict[str, Any], config: ReportConfig) -> str:
        """生成 JSON 報告"""

        logger.info("生成 JSON 報告...")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"evaluation_report_{timestamp}.json"
        filepath = os.path.join(self.output_dir, filename)

        # 確保所有數據都可序列化
        serializable_data = self._make_serializable(report_data)

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(serializable_data, f, ensure_ascii=False, indent=2)

        logger.info(f"JSON 報告已保存至: {filepath}")
        return filepath

    def _generate_excel_report(self, report_data: Dict[str, Any], config: ReportConfig) -> str:
        """生成 Excel 報告"""

        if not OPENPYXL_AVAILABLE:
            logger.error("openpyxl 未安裝，無法生成 Excel 報告")
            return ""

        logger.info("生成 Excel 報告...")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"evaluation_report_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        # 創建 Excel 工作簿
        wb = openpyxl.Workbook()

        # 設定樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 移除默認工作表
        wb.remove(wb.active)

        # 1. 總體指標工作表
        ws_overview = wb.create_sheet("總體指標")

        if report_data["evaluation_results"] and "metrics" in report_data["evaluation_results"]:
            metrics = report_data["evaluation_results"]["metrics"]

            # 標題
            ws_overview["A1"] = "指標"
            ws_overview["B1"] = "數值"

            # 應用標題樣式
            for cell in ["A1", "B1"]:
                ws_overview[cell].font = header_font
                ws_overview[cell].fill = header_fill
                ws_overview[cell].border = border

            # 數據
            metrics_list = [
                ("Precision", metrics.get("precision", 0)),
                ("Recall", metrics.get("recall", 0)),
                ("F1-Score", metrics.get("f1", 0)),
                ("Accuracy", metrics.get("accuracy", 0)),
            ]

            for i, (metric_name, value) in enumerate(metrics_list, 2):
                ws_overview[f"A{i}"] = metric_name
                ws_overview[f"B{i}"] = value
                for cell in [f"A{i}", f"B{i}"]:
                    ws_overview[cell].border = border

            # 調整列寬
            ws_overview.column_dimensions["A"].width = 15
            ws_overview.column_dimensions["B"].width = 15

        # 2. 類別詳細指標工作表
        if (
            report_data["evaluation_results"]
            and "per_class_metrics" in report_data["evaluation_results"]
        ):

            ws_classes = wb.create_sheet("類別指標")
            per_class = report_data["evaluation_results"]["per_class_metrics"]

            # 標題行
            headers = ["類別", "Precision", "Recall", "F1-Score", "Support"]
            for i, header in enumerate(headers, 1):
                cell = ws_classes.cell(row=1, column=i, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            # 數據行
            for row, (class_name, metrics) in enumerate(per_class.items(), 2):
                ws_classes.cell(row=row, column=1, value=class_name).border = border
                ws_classes.cell(row=row, column=2, value=metrics.get("precision", 0)).border = (
                    border
                )
                ws_classes.cell(row=row, column=3, value=metrics.get("recall", 0)).border = border
                ws_classes.cell(row=row, column=4, value=metrics.get("f1-score", 0)).border = border
                ws_classes.cell(row=row, column=5, value=int(metrics.get("support", 0))).border = (
                    border
                )

            # 調整列寬
            for i in range(1, 6):
                ws_classes.column_dimensions[chr(64 + i)].width = 12

        # 3. 錯誤分析工作表
        if report_data["error_analysis"]:
            ws_errors = wb.create_sheet("錯誤分析")

            # 寫入錯誤統計
            ws_errors["A1"] = "錯誤分析統計"
            ws_errors["A1"].font = Font(bold=True, size=14)

            if "statistics" in report_data["error_analysis"]:
                stats = report_data["error_analysis"]["statistics"]

                # 總錯誤數
                ws_errors["A3"] = "總錯誤數"
                ws_errors["B3"] = stats.get("total_errors", 0)

                # 錯誤類型分布
                if "error_types" in stats:
                    row = 5
                    ws_errors[f"A{row}"] = "錯誤類型分布"
                    ws_errors[f"A{row}"].font = Font(bold=True)

                    for error_type, count in stats["error_types"].items():
                        row += 1
                        ws_errors[f"A{row}"] = error_type
                        ws_errors[f"B{row}"] = count

        # 4. 穩健性測試工作表
        if report_data["robustness_results"]:
            ws_robustness = wb.create_sheet("穩健性測試")

            if "summary_statistics" in report_data["robustness_results"]:
                summary = report_data["robustness_results"]["summary_statistics"]

                # 標題
                headers = ["攻擊類型", "成功率", "平均信心下降", "測試數量"]
                for i, header in enumerate(headers, 1):
                    cell = ws_robustness.cell(row=1, column=i, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = border

                # 數據
                for row, (attack_type, stats) in enumerate(summary.items(), 2):
                    ws_robustness.cell(
                        row=row, column=1, value=attack_type.replace("_", " ")
                    ).border = border
                    ws_robustness.cell(
                        row=row, column=2, value=stats.get("attack_success_rate", 0)
                    ).border = border
                    ws_robustness.cell(
                        row=row, column=3, value=stats.get("average_confidence_drop", 0)
                    ).border = border
                    ws_robustness.cell(
                        row=row, column=4, value=stats.get("total_tests", 0)
                    ).border = border

                # 調整列寬
                for i in range(1, 5):
                    ws_robustness.column_dimensions[chr(64 + i)].width = 15

        # 保存文件
        wb.save(filepath)

        logger.info(f"Excel 報告已保存至: {filepath}")
        return filepath

    def _make_serializable(self, obj):
        """轉換對象為可序列化格式"""

        if isinstance(obj, dict):
            return {k: self._make_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._make_serializable(item) for item in obj]
        elif hasattr(obj, "__dict__"):
            return self._make_serializable(obj.__dict__)
        elif isinstance(obj, (np.integer, np.floating)):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif isinstance(obj, (pd.DataFrame, pd.Series)):
            return obj.to_dict()
        else:
            return obj


class HTMLReportGenerator:
    """HTML 專門報告生成器"""

    def __init__(self, template_dir: str = None):
        self.template_dir = template_dir or os.path.join(os.path.dirname(__file__), "templates")

    def generate_interactive_dashboard(self, data: Dict[str, Any]) -> str:
        """生成互動式儀表板"""
        # 實現互動式儀表板生成
        pass


class PDFReportGenerator:
    """PDF 專門報告生成器"""

    def __init__(self, font_path: str = None):
        self.font_path = font_path
        if font_path and REPORTLAB_AVAILABLE:
            try:
                pdfmetrics.registerFont(TTFont("Chinese", font_path))
            except:
                logger.warning("無法註冊中文字體，將使用默認字體")

    def generate_detailed_report(self, data: Dict[str, Any]) -> str:
        """生成詳細的 PDF 報告"""
        # 實現詳細 PDF 報告生成
        pass
